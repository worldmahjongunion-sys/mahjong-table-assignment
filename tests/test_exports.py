import csv
import io

from openpyxl import load_workbook

import exports


def test_build_members_csv_contains_header_and_rows():
    members = [
        {"name": "太郎", "memo": "初段", "created_at": "2026-08-01T00:00:00", "is_active": 1},
        {"name": "次郎", "memo": "", "created_at": "2026-08-02T00:00:00", "is_active": 0},
    ]

    csv_bytes = exports.build_members_csv(members)

    text = csv_bytes.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    assert rows[0] == ["名前", "メモ", "登録日時", "状態"]
    assert rows[1] == ["太郎", "初段", "2026-08-01T00:00:00", "在籍"]
    assert rows[2] == ["次郎", "", "2026-08-02T00:00:00", "引退"]


def test_build_members_csv_empty_list_has_only_header():
    csv_bytes = exports.build_members_csv([])

    text = csv_bytes.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(text)))
    assert rows == [["名前", "メモ", "登録日時", "状態"]]


def test_build_members_xlsx_contains_header_and_rows():
    members = [
        {"name": "太郎", "memo": "初段", "created_at": "2026-08-01T00:00:00", "is_active": 1},
        {"name": "次郎", "memo": None, "created_at": "2026-08-02T00:00:00", "is_active": 0},
    ]

    xlsx_bytes = exports.build_members_xlsx(members)

    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook.active
    rows = [tuple(cell.value for cell in row) for row in sheet.iter_rows()]
    assert rows[0] == ("名前", "メモ", "登録日時", "状態")
    assert rows[1] == ("太郎", "初段", "2026-08-01T00:00:00", "在籍")
    # 空文字セルはxlsx保存→読込を経るとopenpyxl上はNoneに戻る（xlsx形式の仕様。表示上は空欄で正しい）
    assert rows[2] == ("次郎", None, "2026-08-02T00:00:00", "引退")
